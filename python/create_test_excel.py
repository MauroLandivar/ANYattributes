"""Creates a test Excel file that mimics an ANYMARKET Apache POI planilla."""
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active

# Header fills
green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
blue_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")

# Row 1: Column names
headers_row1 = ["ID Anymarket", "Nome del Producto", "Categoria", "SKUS", "Material", "Color", "Peso", "Tamanho"]
for i, h in enumerate(headers_row1, 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.fill = green_fill
    cell.font = Font(bold=True, color="FFFFFFFF")

# Row 2: Marketplace
marketplaces = ["", "", "", "", "FALABELLA", "FALABELLA", "MERCADO_LIVRE", "MERCADO_LIVRE"]
for i, m in enumerate(marketplaces, 1):
    ws.cell(row=2, column=i, value=m).fill = green_fill

# Row 3: Data type
data_types = ["", "", "", "", "Listado", "Texto", "Numero", "Listado"]
for i, dt in enumerate(data_types, 1):
    ws.cell(row=3, column=i, value=dt).fill = green_fill

# Row 4: Dropdown label
labels = ["", "", "", "", "Material del producto", "Color del producto", "Peso en kg", "Tamanho"]
for i, l in enumerate(labels, 1):
    ws.cell(row=4, column=i, value=l).fill = green_fill

# Add data validation for Material (Listado)
from openpyxl.worksheet.datavalidation import DataValidation
dv_material = DataValidation(type="list", formula1='"Algodão,Poliéster,Lana,Seda"', allow_blank=True)
ws.add_data_validation(dv_material)
dv_material.add("E5:E20")

dv_size = DataValidation(type="list", formula1='"PP,P,M,G,GG,XGG"', allow_blank=True)
ws.add_data_validation(dv_size)
dv_size.add("H5:H20")

# Row 5-7: Product data
products = [
    ("ANM001", "Camiseta Básica Algodão Branca", "Vestuário/Camisetas", "SKU-001"),
    ("ANM002", "Calça Jeans Slim Fit", "Vestuário/Calças", "SKU-002"),
    ("ANM003", "Tênis Running Performance", "Calçados/Tênis", "SKU-003"),
]

for row_idx, (pid, name, cat, sku) in enumerate(products, 5):
    ws.cell(row=row_idx, column=1, value=pid)
    ws.cell(row=row_idx, column=2, value=name)
    ws.cell(row=row_idx, column=3, value=cat)
    ws.cell(row=row_idx, column=4, value=sku)

    # E: Material - RED (mandatory, empty)
    ws.cell(row=row_idx, column=5).fill = red_fill

    # F: Color - BLUE (optional, empty)
    ws.cell(row=row_idx, column=6).fill = blue_fill

    # G: Peso - RED (mandatory, empty)
    ws.cell(row=row_idx, column=7).fill = red_fill

    # H: Tamanho - RED (mandatory, empty)
    ws.cell(row=row_idx, column=8).fill = red_fill

test_path = "/tmp/test_anymarket.xlsx"
wb.save(test_path)
print(f"Test file created: {test_path}")
