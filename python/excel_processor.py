#!/usr/bin/env python3
"""
ANYattributes Excel Processor
Reads Apache POI-generated Excel files from ANYMARKET.

File structure:
  Row 1: empty
  Row 2: attribute names (compact, no spaces)
  Row 3: marketplace
  Row 4: data type (Numero / Listado / Texto)
  Row 5: display labels ("ID Anymarket", attribute human labels...)
  Row 6+: product data

Colors (indexed palette):
  indexed=10  → Red  (#FF0000) → mandatory attribute
  indexed=48  → Blue (#3366FF) → optional attribute
"""

import sys
import json
import os
import zipfile
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Color utilities
# ---------------------------------------------------------------------------

# Full Excel indexed color palette (64 entries, BIFF8 standard)
INDEXED_COLORS: dict[int, tuple[int, int, int]] = {
    0:  (0,   0,   0),    # Black
    1:  (255, 255, 255),  # White
    2:  (255, 0,   0),    # Red
    3:  (0,   255, 0),    # Lime
    4:  (0,   0,   255),  # Blue
    5:  (255, 255, 0),    # Yellow
    6:  (255, 0,   255),  # Magenta
    7:  (0,   255, 255),  # Cyan
    8:  (0,   0,   0),    # Black
    9:  (255, 255, 255),  # White
    10: (255, 0,   0),    # Red   ← ANYMARKET mandatory cells
    11: (0,   255, 0),    # Lime
    12: (0,   0,   255),  # Blue
    13: (255, 255, 0),    # Yellow
    14: (255, 0,   255),  # Magenta
    15: (0,   255, 255),  # Cyan
    16: (128, 0,   0),    # Dark Red/Maroon
    17: (0,   128, 0),    # Dark Green
    18: (0,   0,   128),  # Navy
    19: (128, 128, 0),    # Olive
    20: (128, 0,   128),  # Purple
    21: (0,   128, 128),  # Teal
    22: (192, 192, 192),  # Silver
    23: (128, 128, 128),  # Gray
    24: (153, 153, 255),  # Lavender
    25: (153, 51,  102),  # Mauve
    26: (255, 255, 204),  # Light Yellow
    27: (204, 255, 255),  # Light Cyan
    28: (102, 0,   102),  # Dark Purple
    29: (255, 128, 128),  # Salmon
    30: (0,   102, 204),  # Medium Blue
    31: (204, 204, 255),  # Light Lavender
    32: (0,   0,   128),  # Dark Blue
    33: (255, 0,   255),  # Magenta
    34: (255, 255, 0),    # Yellow
    35: (0,   255, 255),  # Cyan
    36: (128, 0,   128),  # Purple
    37: (128, 0,   0),    # Maroon
    38: (0,   128, 128),  # Teal
    39: (0,   0,   255),  # Blue
    40: (0,   204, 255),  # Sky Blue
    41: (204, 255, 255),  # Light Cyan
    42: (204, 255, 204),  # Light Green
    43: (255, 255, 153),  # Light Yellow
    44: (153, 204, 255),  # Light Blue
    45: (255, 153, 204),  # Light Pink
    46: (204, 153, 255),  # Light Purple
    47: (255, 204, 153),  # Peach
    48: (51,  102, 255),  # Cornflower Blue ← ANYMARKET optional cells
    49: (51,  204, 204),  # Teal Blue
    50: (153, 204, 0),    # Yellow-Green
    51: (255, 204, 0),    # Amber
    52: (255, 153, 0),    # Orange
    53: (255, 102, 0),    # Dark Orange
    54: (102, 102, 153),  # Blue-Gray
    55: (150, 150, 150),  # Medium Gray
    56: (0,   51,  102),  # Dark Navy
    57: (51,  153, 102),  # Medium Green
    58: (0,   51,  0),    # Dark Green
    59: (51,  51,  0),    # Dark Olive
    60: (153, 51,  0),    # Brown
    61: (153, 51,  102),  # Dark Mauve
    62: (51,  51,  153),  # Dark Blue-Purple
    63: (51,  51,  51),   # Very Dark Gray
}


def classify_color(r: int, g: int, b: int) -> str | None:
    """Classify RGB as 'red', 'blue', 'green', or None."""
    if r > 160 and g < 120 and b < 120:
        return "red"
    if b > 130 and b > r + 30:
        return "blue"
    if g > 150 and g > r + 30 and g > b + 30:
        return "green"
    return None


def get_cell_color(cell) -> str | None:
    """Read openpyxl cell fill color → 'red', 'blue', 'green', or None."""
    try:
        fill = cell.fill
        if fill is None or fill.fill_type in (None, "none"):
            return None
        fg = fill.fgColor
        if fg.type == "rgb":
            argb = fg.rgb or ""
            if not argb or argb in ("00000000", "FFFFFFFF", "FF000000"):
                return None
            r = int(argb[2:4], 16)
            g = int(argb[4:6], 16)
            b = int(argb[6:8], 16)
            return classify_color(r, g, b)
        elif fg.type == "indexed":
            rgb = INDEXED_COLORS.get(fg.indexed)
            if rgb:
                return classify_color(*rgb)
        elif fg.type == "theme":
            return None  # theme colors need full theme XML resolution
    except Exception:
        pass
    return None


def get_xml_colors(filepath: str) -> dict[str, str]:
    """
    Fallback: parse xlsx zip XML to extract cell colors.
    Returns { "A1": "red", "B3": "blue", ... } for the first sheet.
    """
    result: dict[str, str] = {}
    try:
        with zipfile.ZipFile(filepath) as z:
            names = z.namelist()

            # Parse fills from styles.xml
            styles_xml = z.read("xl/styles.xml")
            styles_root = ET.fromstring(styles_xml)
            ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

            fills: list[str | None] = []
            fills_elem = styles_root.find("x:fills", ns)
            if fills_elem is not None:
                for fill_elem in fills_elem.findall("x:fill", ns):
                    color = None
                    pf = fill_elem.find("x:patternFill", ns)
                    if pf is not None:
                        fg_elem = pf.find("x:fgColor", ns)
                        if fg_elem is not None:
                            argb = fg_elem.get("rgb", "")
                            indexed = fg_elem.get("indexed")
                            if argb and argb not in ("00000000", "FFFFFFFF"):
                                r, g, b = int(argb[2:4], 16), int(argb[4:6], 16), int(argb[6:8], 16)
                                color = classify_color(r, g, b)
                            elif indexed is not None:
                                rgb = INDEXED_COLORS.get(int(indexed))
                                if rgb:
                                    color = classify_color(*rgb)
                    fills.append(color)

            # Map xf index → fill color
            xfs: list[str | None] = []
            cell_xfs = styles_root.find("x:cellXfs", ns)
            if cell_xfs is not None:
                for xf in cell_xfs.findall("x:xf", ns):
                    fill_id = int(xf.get("fillId", 0))
                    apply_fill = xf.get("applyFill", "0")
                    if apply_fill == "1" and fill_id < len(fills):
                        xfs.append(fills[fill_id])
                    else:
                        xfs.append(None)

            # Read first sheet
            sheet_files = sorted(
                n for n in names
                if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
            )
            if sheet_files:
                sheet_xml = z.read(sheet_files[0])
                sheet_root = ET.fromstring(sheet_xml)
                ns2 = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                for c_elem in sheet_root.iter(f"{{{ns2}}}c"):
                    ref = c_elem.get("r", "")
                    s_attr = c_elem.get("s")
                    if s_attr is not None:
                        xf_idx = int(s_attr)
                        if xf_idx < len(xfs) and xfs[xf_idx]:
                            result[ref] = xfs[xf_idx]

    except Exception as e:
        sys.stderr.write(f"[xml_colors] {e}\n")
    return result


# ---------------------------------------------------------------------------
# Data validation helpers
# ---------------------------------------------------------------------------

def extract_dv_options(ws) -> dict[str, list[str]]:
    """Extract dropdown options per column letter from data validations."""
    options: dict[str, list[str]] = {}
    try:
        for dv in ws.data_validations.dataValidation:
            if dv.type != "list" or not dv.formula1:
                continue
            formula = dv.formula1.strip('"').strip("'")
            if "!" in formula:
                continue
            opts = [o.strip() for o in formula.replace(";", ",").split(",") if o.strip()]
            if len(opts) < 2:
                continue
            try:
                for cell_range in dv.sqref.ranges:
                    for col in range(cell_range.min_col, cell_range.max_col + 1):
                        col_letter = get_column_letter(col)
                        if col_letter not in options:
                            options[col_letter] = opts
            except Exception:
                pass
    except Exception:
        pass
    return options


# ---------------------------------------------------------------------------
# Auto-detect file structure
# ---------------------------------------------------------------------------

def detect_structure(ws) -> dict:
    """
    Auto-detect header and data rows.
    Returns { data_start_row, attr_name_row, marketplace_row,
              data_type_row, label_row }
    """
    # Find first row where column A has a real numeric product ID.
    # ANYMARKET product IDs are integers — label rows like "ID Anymarket" must be excluded.
    data_start_row = 6  # safe default for ANYMARKET planillas
    for row in range(3, 20):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        # Accept actual int/float values (openpyxl reads numbers as int/float)
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            data_start_row = row
            break
        # Accept pure-digit strings (e.g. product IDs stored as text)
        s = str(val).strip()
        if s.isdigit():
            data_start_row = row
            break

    label_row = data_start_row - 1      # e.g. row 5
    data_type_row = data_start_row - 2  # e.g. row 4
    marketplace_row = data_start_row - 3  # e.g. row 3
    attr_name_row = data_start_row - 4    # e.g. row 2

    return {
        "data_start_row": data_start_row,
        "attr_name_row": max(1, attr_name_row),
        "marketplace_row": max(1, marketplace_row),
        "data_type_row": max(1, data_type_row),
        "label_row": max(1, label_row),
    }


# ---------------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------------

def analyze_file(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("No active worksheet found")

    max_col = ws.max_column or 0
    max_row = ws.max_row or 0

    # Auto-detect structure
    struct = detect_structure(ws)
    data_start_row = struct["data_start_row"]
    attr_name_row = struct["attr_name_row"]
    marketplace_row = struct["marketplace_row"]
    data_type_row = struct["data_type_row"]
    label_row = struct["label_row"]

    # Build XML fallback color map
    xml_colors = get_xml_colors(filepath)

    def resolve_color(cell) -> str | None:
        color = get_cell_color(cell)
        if color is None:
            ref = f"{get_column_letter(cell.column)}{cell.row}"
            color = xml_colors.get(ref)
        return color

    # Read attribute headers — only columns where attr_name_row has a value
    headers: dict[str, dict] = {}
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        attr_name = ws.cell(row=attr_name_row, column=col).value
        if not attr_name or not str(attr_name).strip():
            continue
        marketplace = ws.cell(row=marketplace_row, column=col).value
        data_type = ws.cell(row=data_type_row, column=col).value
        dropdown_label = ws.cell(row=label_row, column=col).value
        headers[col_letter] = {
            "name": str(attr_name).strip(),
            "marketplace": str(marketplace).strip() if marketplace else "",
            "data_type": str(data_type).strip() if data_type else "",
            "dropdown_label": str(dropdown_label).strip() if dropdown_label else "",
        }

    # Extract data validation options
    dv_options = extract_dv_options(ws)

    # Parse dropdown labels as fallback options
    label_options: dict[str, list[str]] = {}
    for col_letter, hdr in headers.items():
        label = hdr.get("dropdown_label", "")
        if label and (";" in label or "," in label):
            opts = [o.strip() for o in label.replace(";", ",").split(",") if o.strip()]
            if len(opts) > 1:
                label_options[col_letter] = opts

    def get_options(col_letter: str) -> list[str]:
        return dv_options.get(col_letter) or label_options.get(col_letter) or []

    # Scan product rows
    cells_to_fill: list[dict] = []
    total_products = 0
    red_cells_empty = 0
    blue_cells_empty = 0
    unknown_cells_empty = 0

    for row in range(data_start_row, max_row + 1):
        product_id = ws.cell(row=row, column=1).value
        if product_id is None or str(product_id).strip() == "":
            continue

        total_products += 1
        product_name = ws.cell(row=row, column=2).value
        category = ws.cell(row=row, column=3).value
        skus = ws.cell(row=row, column=4).value

        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            if col_letter not in headers:
                continue

            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            if cell_value is not None and str(cell_value).strip() != "":
                continue  # already has a value

            color = resolve_color(cell)

            if color == "red":
                red_cells_empty += 1
            elif color == "blue":
                blue_cells_empty += 1
            else:
                unknown_cells_empty += 1
                color = "none"

            hdr = headers[col_letter]
            cells_to_fill.append({
                "row": row,
                "col": col,
                "col_letter": col_letter,
                "product_id": str(product_id),
                "product_name": str(product_name) if product_name else "",
                "category": str(category) if category else "",
                "skus": str(skus) if skus else "",
                "attribute_name": hdr["name"],
                "marketplace": hdr["marketplace"],
                "data_type": hdr["data_type"],
                "dropdown_label": hdr["dropdown_label"],
                "options": get_options(col_letter),
                "color": color,
            })

    wb.close()

    return {
        "total_products": total_products,
        "red_cells_empty": red_cells_empty,
        "blue_cells_empty": blue_cells_empty,
        "unknown_cells_empty": unknown_cells_empty,
        "headers": headers,
        "cells": cells_to_fill,
        "filename": os.path.basename(filepath),
        "structure": struct,
    }


# ---------------------------------------------------------------------------
# Write filled values
# ---------------------------------------------------------------------------

def write_file(source_filepath: str, cells_data: list[dict], output_filepath: str) -> None:
    wb = openpyxl.load_workbook(source_filepath)
    ws = wb.active

    if ws is None:
        raise ValueError("No active worksheet found")

    for item in cells_data:
        row = item["row"]
        col = item["col"]
        value = item.get("value")
        if value is None or str(value).strip() == "":
            continue

        cell = ws.cell(row=row, column=col)

        data_type = item.get("data_type", "").lower()
        if "numero" in data_type or "number" in data_type:
            try:
                cell.value = float(str(value).replace(",", "."))
            except (ValueError, TypeError):
                cell.value = value
        else:
            cell.value = str(value)

        existing_font = cell.font
        cell.font = Font(
            name=existing_font.name,
            size=existing_font.size,
            bold=existing_font.bold,
            italic=existing_font.italic,
            underline=existing_font.underline,
            strike=existing_font.strike,
            color="FFFFFFFF",  # white = AI-filled indicator
        )

    wb.save(output_filepath)
    wb.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: excel_processor.py <analyze|write> [args]"}))
        sys.exit(1)

    command = sys.argv[1]

    if command == "analyze":
        if len(sys.argv) < 3:
            print(json.dumps({"error": "Usage: analyze <filepath>"}))
            sys.exit(1)
        fp = sys.argv[2]
        if not os.path.exists(fp):
            print(json.dumps({"error": f"File not found: {fp}"}))
            sys.exit(1)
        try:
            print(json.dumps(analyze_file(fp), ensure_ascii=False))
        except Exception as e:
            print(json.dumps({"error": str(e)}))
            sys.exit(1)

    elif command == "write":
        if len(sys.argv) < 5:
            print(json.dumps({"error": "Usage: write <source> <cells_json_file> <output>"}))
            sys.exit(1)
        source, cells_json_path, output = sys.argv[2], sys.argv[3], sys.argv[4]
        try:
            with open(cells_json_path, encoding="utf-8") as f:
                cells_data = json.load(f)
            write_file(source, cells_data, output)
            print(json.dumps({"success": True, "output": output}))
        except Exception as e:
            print(json.dumps({"error": str(e)}))
            sys.exit(1)

    else:
        print(json.dumps({"error": f"Unknown command: {command}"}))
        sys.exit(1)
